#!/usr/bin/env python3
"""
生成"焦点球星"的 7m ID 填写 Excel
- 每行一个球员，带 7m 搜索超链接
- 手动填写 7m 球员 ID 后，运行 fifa_batch_scrape.py 批量抓取
用法:
  python3 gen_star_xlsx.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import urllib.parse
import os

SRC_XLSX = '2026世界杯观赛指南.xlsx'
OUT_XLSX = '焦点球星_7m_ID填写.xlsx'
DEST_SHEET = '焦点球星_7m_ID填写'


def make_thin_border():
    s = Side(style='thin', color='BBBBBB')
    return Border(left=s, right=s, top=s, bottom=s)


def build_xlsx():
    wb_src = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws_src = wb_src['焦点球星']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(DEST_SHEET)

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='234099', end_color='234099', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    link_font = Font(color='0563C1', underline='single', size=10)
    thin = make_thin_border()

    # ── 表头 ──────────────────────────────────
    headers = [
        ('#', 5),
        ('姓名', 14),
        ('国籍', 10),
        ('位置', 10),
        ('年龄', 6),
        ('效力俱乐部', 22),
        ('看点', 40),
        ('7m 搜索', 12),
        ('7m 球员ID\n(手动填写)', 16),
        ('操作指引', 30),
    ]
    for col, (title, _) in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin
    ws.row_dimensions[1].height = 30

    # ── 数据行 ────────────────────────────────
    for row in range(2, ws_src.max_row + 1):
        name = ws_src.cell(row=row, column=1).value or ''
        nationality = ws_src.cell(row=row, column=2).value or ''
        pos = ws_src.cell(row=row, column=3).value or ''
        age = ws_src.cell(row=row, column=4).value or ''
        club = ws_src.cell(row=row, column=5).value or ''
        note = ws_src.cell(row=row, column=6).value or ''

        r = row  # Excel 行号（源文件行号 = 目标文件行号，因为都有表头）

        # 序号
        c = ws.cell(row=r, column=1, value=r - 1)
        c.alignment = center
        c.border = thin

        # 姓名
        c = ws.cell(row=r, column=2, value=str(name))
        c.alignment = left
        c.border = thin

        # 国籍
        c = ws.cell(row=r, column=3, value=str(nationality))
        c.alignment = center
        c.border = thin

        # 位置
        c = ws.cell(row=r, column=4, value=str(pos))
        c.alignment = center
        c.border = thin

        # 年龄
        c = ws.cell(row=r, column=5, value=age)
        c.alignment = center
        c.border = thin

        # 俱乐部
        c = ws.cell(row=r, column=6, value=str(club))
        c.alignment = left
        c.border = thin

        # 看点
        c = ws.cell(row=r, column=7, value=str(note))
        c.alignment = left
        c.border = thin

        # 7m 搜索超链接
        q = urllib.parse.quote(str(name))
        url = f'https://search.7m.com.cn/search.aspx?q={q}&type=player&lang=gb'
        c = ws.cell(row=r, column=8)
        c.value = f'=HYPERLINK("{url}","🔍 搜索")'
        c.font = link_font
        c.alignment = center
        c.border = thin

        # 7m 球员 ID（待手动填写）
        c = ws.cell(row=r, column=9, value='')
        c.alignment = center
        c.border = thin

        # 操作指引
        c = ws.cell(row=r, column=10,
                      value='① 点左侧"🔍 搜索" → 新标签页打开 7m\n'
                              '② 找到对应该球员的页面，点击进入\n'
                              '③ 从 URL 中复制数字 ID（如 97901）\n'
                              '④ 粘贴到左侧"I"列')
        c.alignment = left
        c.border = thin

        ws.row_dimensions[r].height = 40

    # ── 列宽 ──────────────────────────────────
    widths = [5, 14, 10, 10, 6, 22, 40, 12, 16, 30]
    for col, w_ in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w_

    wb.save(OUT_XLSX)
    print(f'[完成] 已生成: {OUT_XLSX}')
    print(f'  文件位置: {os.path.abspath(OUT_XLSX)}')
    print(f'  共 {ws_src.max_row - 1} 名焦点球星')
    print()
    print('  使用步骤：')
    print('  1. 用 WPS/Excel 打开此文件')
    print('  2. 点击"🔍 搜索"列，浏览器打开 7m 搜索结果')
    print('  3. 找到该球员的 7m 页面，从 URL 中复制球员 ID')
    print('  4. 将 ID 填入"I"列')
    print('  5. 全部填完后，运行: python fifa_batch_scrape.py filled.xlsx')


if __name__ == '__main__':
    if not os.path.exists(SRC_XLSX):
        print(f'[错误] 找不到 {SRC_XLSX}')
        exit(1)
    build_xlsx()
