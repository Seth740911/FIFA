#!/usr/bin/env python3
"""
读取填好 7m ID 的 Excel/CSV，批量抓取球员数据并导出 Excel。
用法:
  python fifa_batch_scrape.py 焦点球星_7m_ID填写.xlsx
  python fifa_batch_scrape.py filled.csv
"""

import json
import sys
import os
import subprocess
import csv
import argparse

SCRAPE_PY = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'scrape_7m_player.py')
TMP_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_tmp_players.json')


def read_filled_excel(xlsx_path):
    """读取填好7m ID的Excel文件"""
    try:
        import openpyxl
    except ImportError:
        print('[错误] 需要 openpyxl，请运行: pip install openpyxl', file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    players = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value or ''   # 姓名
        pid = str(ws.cell(row=row, column=9).value or '').strip()  # 7m ID
        if pid and pid.isdigit():
            players.append({'name': name, '7m_id': pid})
    return players


def read_filled_csv(csv_path):
    """读取填好7m ID的CSV文件"""
    players = []
    with open(csv_path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            pid = row.get('7m球员ID', row.get('7m ID', '')).strip()
            name = row.get('姓名', row.get('name_cn', ''))
            if pid and pid.isdigit():
                players.append({'name': name, '7m_id': pid})
    return players


def batch_scrape(players):
    """调用 scrape_7m_player.py 批量抓取"""
    if not players:
        print('[错误] 没有有效的7m球员ID', file=sys.stderr)
        sys.exit(1)

    ids = [p['7m_id'] for p in players]
    print(f'[抓取] 共 {len(ids)} 名球员')
    print(f'  IDs: {ids[:5]}{"..." if len(ids) > 5 else ""}')

    cmd = [sys.executable, SCRAPE_PY] + ids + ['-o', TMP_JSON]
    print(f'[执行] {" ".join(cmd)}')
    ret = subprocess.run(cmd, cwd=os.path.dirname(os.path.abspath(__file__)))

    if ret.returncode != 0 or not os.path.exists(TMP_JSON):
        print('[错误] 抓取失败', file=sys.stderr)
        sys.exit(1)

    with open(TMP_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)

    results = data if isinstance(data, list) else [data]

    # 导出Excel
    import importlib.util
    spec = importlib.util.spec_from_file_location('scrape_mod', SCRAPE_PY)
    scrape_mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(scrape_mod)

    output = '2026世界杯球员数据_7m.xlsx'
    scrape_mod.export_to_xlsx(results, output)
    os.remove(TMP_JSON)
    print(f'[完成] Excel已保存: {output}')
    return output


def main():
    parser = argparse.ArgumentParser(description='批量抓取7m球员数据')
    parser.add_argument('input', help='填好7m ID的Excel或CSV文件')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f'[错误] 找不到文件: {args.input}', file=sys.stderr)
        sys.exit(1)

    if args.input.endswith('.xlsx'):
        players = read_filled_excel(args.input)
    elif args.input.endswith('.csv'):
        players = read_filled_csv(args.input)
    else:
        print('[错误] 只支持 .xlsx 或 .csv 文件', file=sys.stderr)
        sys.exit(1)

    if not players:
        print('[错误] 文件中没有有效的7m球员ID', file=sys.stderr)
        print('  请确保在Excel的"7m 球员ID(手动填写)"列填写了数字ID', file=sys.stderr)
        sys.exit(1)

    print(f'[信息] 读取到 {len(players)} 名球员:')
    for p in players[:5]:
        print(f'  {p["name"]} -> ID: {p["7m_id"]}')
    if len(players) > 5:
        print(f'  ... 共 {len(players)} 人')

    batch_scrape(players)


if __name__ == '__main__':
    main()
