#!/usr/bin/env python3
"""
7m体育网 球员数据爬虫 + Excel导出
用法:
  python scrape_7m_player.py 97901
  python scrape_7m_player.py 97901 -o 97901.json
  python scrape_7m_player.py 97901 45 -x -o players.xlsx
  python scrape_7m_player.py -i ids.txt -x -o players.xlsx
"""

import json
import re
import sys
import time
import argparse
from urllib.request import Request, urlopen

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
}

_LINEUP_ARR = None
_PLAYER_DATA_TITLE = None


def http_get(url, referer=None, timeout=10):
    headers = dict(HEADERS)
    if referer:
        headers['Referer'] = referer
    req = Request(url, headers=headers)
    with urlopen(req, timeout=timeout) as r:
        return r.read().decode('utf-8', errors='ignore')


def load_constants(lang='gb'):
    global _LINEUP_ARR, _PLAYER_DATA_TITLE
    if _LINEUP_ARR is not None:
        return
    try:
        js = http_get(f'https://static.7m.com.cn/js/player/const/{lang}.js')
        m = re.search(r'LINEUP_ARR\s*=\s*(\[.*?\])\s*;', js, re.DOTALL)
        if m:
            _LINEUP_ARR = json.loads(m.group(1))
        m2 = re.search(r'PLAYER_DATA_TITLE\s*=\s*(\[.*?\])\s*;', js, re.DOTALL)
        if m2:
            _PLAYER_DATA_TITLE = json.loads(m2.group(1))
    except Exception as e:
        print(f'  [警告] 加载常量失败: {e}', file=sys.stderr)
        _LINEUP_ARR = ["前锋", "中场", "后卫", "守门员", "", "", "", "", "", "其它"]
        _PLAYER_DATA_TITLE = ["国籍", "生日", "生物周期", "身高", "体重",
                              "效力球队", "场上位置", "加盟日期", "转会费",
                              "前度效力球队", "曾经效力球队", "俱乐部球衣", "球员身价"]


def extract_json_var(js_text, var_name):
    prefix = f'var {var_name} ='
    idx = js_text.find(prefix)
    if idx < 0:
        return None
    start = idx + len(prefix)
    while start < len(js_text) and js_text[start] in ' \t\r\n':
        start += 1
    if start >= len(js_text) or js_text[start] != '{':
        return None
    depth = 0
    in_str = False
    str_ch = None
    i = start
    while i < len(js_text):
        c = js_text[i]
        if in_str:
            if c == '\\' and i + 1 < len(js_text):
                i += 1
            elif c == str_ch:
                in_str = False
        else:
            if c in ('"', "'"):
                in_str = True
                str_ch = c
            elif c == '{':
                depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0:
                    return json.loads(js_text[start:i + 1])
        i += 1
    return None


def parse_info_fun_js(js):
    mapping = {}
    m = re.search(r'var playerName = playerInfo\["(.*?)"\]', js)
    if m:
        mapping[m.group(1)] = 'name_cn'
    m2 = re.search(r'playerName \+=.*?playerInfo\["(.*?)"\]', js)
    if m2:
        mapping[m2.group(1)] = 'name_en'
    title_field = {
        0: 'nationality',
        1: 'birthday',
        3: 'height',
        4: 'weight',
        5: 'club',
        6: 'position',
        7: 'join_date',
        8: 'transfer_fee',
        9: 'former_club',
        10: 'once_club',
        11: 'shirt_no',
        12: 'market_value',
    }
    positions = [(m.start(), int(m.group(1)))
                 for m in re.finditer(r'PLAYER_DATA_TITLE\[(\d+)\]', js)]
    positions.sort()
    key_occs = [(m.start(), m.group(1))
                  for m in re.finditer(r'playerInfo\["(.*?)"\]', js)]
    for pos, tidx in positions:
        if tidx not in title_field:
            continue
        best = None
        best_pos = -1
        for kpos, k in key_occs:
            if kpos < pos and kpos > best_pos and k not in mapping:
                best = k
                best_pos = kpos
        if best:
            mapping[best] = title_field[tidx]
    for km in re.finditer(r'playerInfo\["(.*?)"\]', js):
        k = km.group(1)
        if k in mapping:
            continue
        rest = js[km.end():km.end() + 200]
        if 'profile_td' in rest:
            mapping[k] = 'profile'
        elif 'golry_td' in rest:
            mapping[k] = 'honours'
    return mapping


def fetch_player_info(pid, lang='gb'):
    load_constants(lang)
    nc = str(int(time.time() * 1000))
    url = f'https://player.7m.com.cn/v2/encrypt/fun/getinfo.php?id={pid}&lang={lang}&nc={nc}'
    referer = f'https://player.7m.com.cn/{pid}/index_{lang}.shtml'
    js = http_get(url, referer=referer)
    info_enc = extract_json_var(js, 'playerInfo')
    if not info_enc:
        return None, f'无法解析playerInfo，响应前300字: {js[:300]}'
    eindex = info_enc.get('e_index', '')
    field_map = {}
    if eindex:
        try:
            js_fun = http_get(
                f'https://player.7m.com.cn/v2/encrypt/fun/getinfofun.php?eindex={eindex}&lang={lang}'
            )
            field_map = parse_info_fun_js(js_fun)
        except Exception as e:
            print(f'  [警告] 获取字段映射失败: {e}', file=sys.stderr)
    result = {}
    for k, v in info_enc.items():
        if k in ('e_index', 'link'):
            result[k] = v
        elif k in field_map:
            result[field_map[k]] = v
        else:
            guessed = _guess_field(k, v, result)
            if guessed:
                result[guessed] = v
            else:
                result[f'_raw_{k}'] = v
    if 'position' in result and _LINEUP_ARR:
        try:
            idx = int(result['position'])
            if 0 <= idx < len(_LINEUP_ARR):
                result['position'] = _LINEUP_ARR[idx]
        except (ValueError, TypeError):
            pass
    if 'birthday' in result:
        result['birthday'] = result['birthday'].replace(',', '-').strip()
    return result, None


def _guess_field(k, v, result_so_far):
    if not isinstance(v, str):
        return None
    s = v.strip()
    if re.match(r'^\d{4}[-,]\d{1,2}[-,]\d{1,2}$', s):
        return 'birthday' if 'birthday' not in result_so_far else None
    if re.match(r'^\d+cm$', s):
        return 'height' if 'height' not in result_so_far else None
    if re.match(r'^\d+kg$', s):
        return 'weight' if 'weight' not in result_so_far else None
    return None


def _parse_vs(vs_str):
    parts = vs_str.split(',')
    return {
        'competition_id': parts[1] if len(parts) > 1 else '',
        'team1_id': parts[2] if len(parts) > 2 else '',
        'team2_id': parts[3] if len(parts) > 3 else '',
        'player_team_id': parts[4] if len(parts) > 4 else '',
        'venue_flag': parts[5] if len(parts) > 5 else '',
        'score1': parts[6] if len(parts) > 6 else '',
        'score2': parts[7] if len(parts) > 7 else '',
    }


def _parse_s(s_str):
    parts = s_str.split(',')
    return {
        'goals': parts[0] if len(parts) > 0 else '',
        'penalties': parts[1] if len(parts) > 1 else '',
        'own_goals': parts[2] if len(parts) > 2 else '',
        'yellow_cards': parts[3] if len(parts) > 3 else '',
        'red_cards': parts[4] if len(parts) > 4 else '',
        'minutes': parts[5] if len(parts) > 5 else '',
    }


def decrypt_stats(stats_enc, lang='gb'):
    if not stats_enc:
        return None
    result = {}
    for k, v in stats_enc.items():
        if k in ('e_index', 'link'):
            result[k] = v
        elif isinstance(v, dict):
            first_val = next(iter(v.values()), '')
            if isinstance(first_val, dict) and 'n' in first_val:
                result['competitions'] = v
            else:
                result['teams'] = v
        elif isinstance(v, list):
            matches = []
            for rec in v:
                vs_info = _parse_vs(rec.get('vs', ''))
                s_info = _parse_s(rec.get('s', ''))
                date_str = rec.get('t', '').replace(',', '-')
                matches.append({
                    'match_id': rec.get('vs', '').split(',')[0] if rec.get('vs') else '',
                    'date': date_str,
                    **vs_info,
                    **s_info,
                })
            result['matches'] = matches
    return result


def fetch_player_stats(pid, lang='gb'):
    nc = str(int(time.time() * 1000))
    url = f'https://player.7m.com.cn/v2/encrypt/fun/getstats.php?id={pid}&lang={lang}&nc={nc}'
    referer = f'https://player.7m.com.cn/{pid}/index_{lang}.shtml'
    try:
        js = http_get(url, referer=referer, timeout=15)
        stats_enc = extract_json_var(js, 'playerStats')
        return decrypt_stats(stats_enc, lang) if stats_enc else None
    except Exception as e:
        print(f'  [警告] 获取stats失败: {e}', file=sys.stderr)
        return None


def fetch_intime(pid):
    url = f'https://player.7m.com.cn/{pid}/data/intime.js'
    try:
        js = http_get(url, timeout=10)
        return extract_json_var(js, 'intime')
    except Exception:
        return None


def scrape_player(pid, lang='gb'):
    result = {'player_id': pid}
    info, err = fetch_player_info(pid, lang)
    if err:
        result['error'] = err
        return result
    result['info'] = info
    stats = fetch_player_stats(pid, lang)
    if stats:
        result['stats'] = stats
    intime = fetch_intime(pid)
    if intime:
        result['intime'] = intime
    return result


# ─── Excel 导出 ───────────────────────────────────────────────────────────────

INFO_COLUMNS = [
    ('球员ID', 'player_id'),
    ('中文名', 'name_cn'),
    ('英文名', 'name_en'),
    ('国籍', 'nationality'),
    ('生日', 'birthday'),
    ('身高', 'height'),
    ('体重', 'weight'),
    ('俱乐部', 'club'),
    ('位置', 'position'),
    ('球衣号码', 'shirt_no'),
    ('加盟日期', 'join_date'),
    ('转会费', 'transfer_fee'),
    ('前度效力球队', 'former_club'),
    ('曾经效力球队', 'once_club'),
    ('球员身价', 'market_value'),
]

MATCH_COLUMNS = [
    ('球员ID', 'player_id'),
    ('中文名', '_name_cn'),
    ('日期', 'date'),
    ('赛事ID', 'competition_id'),
    ('主队ID', 'team1_id'),
    ('客队ID', 'team2_id'),
    ('主队进球', 'score1'),
    ('客队进球', 'score2'),
    ('进球', 'goals'),
    ('点球', 'penalties'),
    ('乌龙球', 'own_goals'),
    ('黄牌', 'yellow_cards'),
    ('红牌', 'red_cards'),
    ('上场时间(分钟)', 'minutes'),
]


def _get_info(data, key):
    return data.get('info', {}).get(key, '')


def export_to_xlsx(all_results, xlsx_path):
    """
    将抓取结果导出为 Excel（.xlsx），兼容 WPS。
    三个 Sheet：
      1. 基本信息  - 每个球员一行
      2. 比赛统计  - 每个球员每场比赛一行
      3. 荣誉简介  - 球员荣誉和简介
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('[错误] 需要 openpyxl 库，请运行: pip install openpyxl', file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    wrap_align = Alignment(wrap_text=True, vertical='top', horizontal='left')
    center_align = Alignment(horizontal='center', vertical='center')

    # ── Sheet 1: 基本信息 ──
    ws = wb.create_sheet('基本信息')
    for col, (title, _) in enumerate(INFO_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws.row_dimensions[1].height = 20

    for row_idx, data in enumerate(all_results, 2):
        info = data.get('info', {})
        for col, (_, key) in enumerate(INFO_COLUMNS, 1):
            val = info.get(key, '')
            cell = ws.cell(row=row_idx, column=col, value=str(val) if val else '')
            cell.alignment = wrap_align
        ws.row_dimensions[row_idx].height = 16

    for col_idx in range(1, len(INFO_COLUMNS) + 1):
        title = INFO_COLUMNS[col_idx - 1][0]
        max_len = len(title) * 2
        for row_idx in range(2, len(all_results) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value or ''
            char_len = 0
            for c in str(val):
                char_len += 2 if '\u4e00' <= c <= '\u9fff' else 1
            max_len = max(max_len, char_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # ── Sheet 2: 比赛统计 ──
    ws2 = wb.create_sheet('比赛统计')
    for col, (title, _) in enumerate(MATCH_COLUMNS, 1):
        cell = ws2.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws2.row_dimensions[1].height = 20

    match_row = 2
    for data in all_results:
        pid = data.get('player_id', '')
        name_cn = _get_info(data, 'name_cn')
        matches = data.get('stats', {}).get('matches', [])
        if not matches:
            ws2.cell(row=match_row, column=1, value=pid)
            ws2.cell(row=match_row, column=2, value=name_cn)
            match_row += 1
        else:
            for m in matches:
                for col, (_, key) in enumerate(MATCH_COLUMNS, 1):
                    if key == 'player_id':
                        val = pid
                    elif key == '_name_cn':
                        val = name_cn
                    else:
                        val = m.get(key, '')
                    ws2.cell(row=match_row, column=col, value=str(val) if val else '')
                match_row += 1

    for row_idx in range(2, match_row):
        ws2.row_dimensions[row_idx].height = 16
    for col_idx in range(1, len(MATCH_COLUMNS) + 1):
        title = MATCH_COLUMNS[col_idx - 1][0]
        max_len = len(title) * 2
        for row_idx in range(2, match_row):
            val = ws2.cell(row=row_idx, column=col_idx).value or ''
            char_len = 0
            for c in str(val):
                char_len += 2 if '\u4e00' <= c <= '\u9fff' else 1
            max_len = max(max_len, char_len)
        ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    # ── Sheet 3: 荣誉简介 ──
    ws3 = wb.create_sheet('荣誉简介')
    for col, title in enumerate(['球员ID', '中文名', '个人简介', '荣誉'], 1):
        cell = ws3.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws3.row_dimensions[1].height = 20

    for row_idx, data in enumerate(all_results, 2):
        info = data.get('info', {})
        ws3.cell(row=row_idx, column=1, value=data.get('player_id', ''))
        ws3.cell(row=row_idx, column=2, value=info.get('name_cn', ''))
        profile = re.sub(r'<[^>]+>', '', info.get('profile', ''))
        honours = re.sub(r'<[^>]+>', '', info.get('honours', ''))
        # 把 &quot; 等HTML实体转回来
        for orig, repl in [('&quot;', '"'), ('&amp;', '&'), ('&lt;', '<'), ('&gt;', '>'), ('&#8212;', '—')]:
            profile = profile.replace(orig, repl)
            honours = honours.replace(orig, repl)
        profile_cell = ws3.cell(row=row_idx, column=3, value=profile)
        profile_cell.alignment = wrap_align
        honours_cell = ws3.cell(row=row_idx, column=4, value=honours)
        honours_cell.alignment = wrap_align
        ws3.row_dimensions[row_idx].height = 60

    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 40
    ws3.column_dimensions['D'].width = 50

    wb.save(xlsx_path)
    print(f'[完成] Excel已保存到 {xlsx_path}', file=sys.stderr)


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='7m体育网 球员数据爬虫')
    parser.add_argument('pids', nargs='*', help='球员ID列表')
    parser.add_argument('-i', '--input', help='从文件读取球员ID（每行一个）')
    parser.add_argument('-o', '--output', default='-', help='输出文件（.json 或 .xlsx）')
    parser.add_argument('-x', '--xlsx', action='store_true', help='导出为Excel格式')
    parser.add_argument('-l', '--lang', default='gb', help='语言（gb=简体中文）')
    args = parser.parse_args()

    pids = list(args.pids)
    if args.input:
        with open(args.input, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    pids.append(line)

    if not pids:
        parser.print_help()
        sys.exit(1)

    all_results = []
    for pid in pids:
        print(f'[抓取中] 球员ID: {pid}', file=sys.stderr)
        data = scrape_player(pid, args.lang)
        all_results.append(data)
        if len(pids) > 1:
            time.sleep(0.5)

    is_xlsx = args.xlsx or (args.output != '-' and args.output.endswith('.xlsx'))

    if is_xlsx:
        xlsx_path = args.output if args.output != '-' else 'players.xlsx'
        export_to_xlsx(all_results, xlsx_path)
    else:
        output = all_results[0] if len(all_results) == 1 else all_results
        json_str = json.dumps(output, ensure_ascii=False, indent=2)
        if args.output != '-':
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(json_str)
            print(f'[完成] 已保存到 {args.output}', file=sys.stderr)
        else:
            print(json_str)


if __name__ == '__main__':
    main()
