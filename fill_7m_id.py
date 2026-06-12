#!/usr/bin/env python3
"""
生成 2026世界杯焦点球星的 7m ID 填写助手
- 生成带超链接的 HTML 页面，点击直接跳转到 7m 搜索结果
- 填写 ID 后，点击"导出 CSV"按钮，自动下载 CSV
- 然后用 fifa_build_id_map.py 批量抓取
"""

import openpyxl
import os

EXCEL_PATH = '2026世界杯观赛指南.xlsx'
OUTPUT_HTML = 'fill_7m_id.html'


def read_star_players(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['焦点球星']
    players = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        players.append({
            'name_cn': str(name).strip(),
            'nationality': ws.cell(row=row, column=2).value or '',
            'position': ws.cell(row=row, column=3).value or '',
            'age': ws.cell(row=row, column=4).value or '',
            'club': ws.cell(row=row, column=5).value or '',
            'note': ws.cell(row=row, column=6).value or '',
        })
    return players


def generate_html(players, output_path):
    rows_html = ''
    for i, p in enumerate(players):
        search_url = f"https://search.7m.com.cn/search.aspx?q={p['name_cn']}&type=player&lang=gb"
        rows_html += f'''
        <tr>
            <td>{i+1}</td>
            <td>{p['name_cn']}</td>
            <td>{p['nationality']}</td>
            <td>{p['position']}</td>
            <td>{p['club']}</td>
            <td><a href="{search_url}" target="_blank">🔍 搜索</a></td>
            <td><input type="text" id="id_{i}" placeholder="填7m ID" style="width:80px"></td>
            <td><button onclick="autoFill({i})">✅ 确认</button></td>
            <td id="status_{i}"></td>
        </tr>
'''

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>2026世界杯焦点球星 - 7m ID 填写助手</title>
<style>
    body {{ font-family: "Microsoft YaHei", sans-serif; margin: 20px; background: #f5f5f5; }}
    h1 {{ color: #333; }}
    .tip {{ background: #fff3cd; padding: 12px; border-radius: 6px; margin-bottom: 16px; }}
    table {{ border-collapse: collapse; width: 100%; background: white; }}
    th {{ background: #4472C4; color: white; padding: 8px; position: sticky; top: 0; }}
    td {{ padding: 6px 10px; border-bottom: 1px solid #ddd; }}
    tr:hover {{ background: #f0f4ff; }}
    input {{ padding: 4px; border: 1px solid #ccc; border-radius: 4px; }}
    button {{ padding: 4px 10px; border: none; border-radius: 4px; cursor: pointer; }}
    .btn-search {{ background: #28a745; color: white; }}
    .btn-confirm {{ background: #007bff; color: white; }}
    .status-ok {{ color: #28a745; font-weight: bold; }}
    .toolbar {{ margin: 12px 0; }}
    .toolbar button {{ padding: 8px 16px; margin-right: 8px; font-size: 14px; }}
    a {{ color: #007bff; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
</style>
</head>
<body>

<h1>⚽ 2026世界杯焦点球星 - 7m ID 填写助手</h1>

<div class="tip">
    <b>使用说明：</b><br>
    1. 点击"🔍 搜索"链接，在浏览器新标签页找到该球员的 7m 页面<br>
    2. 从 URL 中复制球员 ID（格式：<code>https://player.7m.com.cn/<b>XXXXXX</b>/</code>）<br>
    3. 将 ID 填入"7m ID"列，点击"✅ 确认"<br>
    4. 全部填完后，点击下方"📥 导出 CSV"按钮
</div>

<div class="toolbar">
    <button onclick="exportCSV()" style="background:#28a745;color:white;">📥 导出 CSV</button>
    <button onclick="exportExcel()" style="background:#007bff;color:white;">📊 导出 Excel</button>
    <span id="progress"></span>
</div>

<table>
    <thead>
        <tr>
            <th>#</th>
            <th>姓名</th>
            <th>国籍</th>
            <th>位置</th>
            <th>效力俱乐部</th>
            <th>操作</th>
            <th>7m ID</th>
            <th>确认</th>
            <th>状态</th>
        </tr>
    </thead>
    <tbody>
        {rows_html}
    </tbody>
</table>

<p style="margin-top:20px; color:#666; font-size:12px;">
    ⚠️ 提示：导出 CSV 后，运行 <code>python fifa_build_id_map.py filled.csv</code> 进行批量抓取
</p>

<script>
const players = {str([{'name': p['name_cn'], 'nationality': p['nationality'], 'position': p['position'], 'club': p['club']} for p in players]).replace("'", '"')};

function autoFill(idx) {{
    const id = document.getElementById('id_' + idx).value.trim();
    if (!id || !/^\d+$/.test(id)) {{
        alert('请输入有效的数字 ID');
        return;
    }}
    document.getElementById('status_' + idx).innerHTML = '<span class="status-ok">✅ 已确认</span>';
    updateProgress();
}}

function updateProgress() {{
    const total = {len(players)};
    let filled = 0;
    for (let i = 0; i < total; i++) {{
        const val = document.getElementById('id_' + i).value.trim();
        if (val && /^\d+$/.test(val)) filled++;
    }}
    document.getElementById('progress').innerText = `进度: ${{filled}}/${{total}}`;
}}

function exportCSV() {{
    let csv = '\ufeff分组,队伍,号码,姓名,英文名,位置,出生日期,7m球员ID\n';
    // 这里只有焦点球星，分组/队伍信息需要从原始Excel获取
    // 简化版：只导出 姓名 + 7m ID
    let lines = [];
    for (let i = 0; i < {len(players)}; i++) {{
        const id = document.getElementById('id_' + i).value.trim();
        if (id) {{
            lines.push(`},${{players[i]['name_cn']}},${{id}}`);
        }}
    }}
    if (lines.length === 0) {{
        alert('请先填写至少一个球员的 7m ID');
        return;
    }}
    csv = '\ufeff姓名,7m球员ID\n' + lines.join('\n');
    download(csv, 'star_players_7m_id.csv', 'text/csv;charset=utf-8');
}}

function exportExcel() {{
    // 生成一个简单的 HTML table，用户可以复制粘贴到 Excel
    let html = '<table border="1"><tr><th>姓名</th><th>国籍</th><th>位置</th><th>俱乐部</th><th>7m ID</th></tr>';
    for (let i = 0; i < {len(players)}; i++) {{
        const id = document.getElementById('id_' + i).value.trim();
        if (id) {{
            html += `<tr><td>${{players[i]['name_cn']}}</td><td>${{players[i]['nationality']}}</td><td>${{players[i]['position']}}</td><td>${{players[i]['club']}}</td><td>${{id}}</td></tr>`;
        }}
    }}
    html += '</table>';
    const w = window.open('');
    w.document.write(html);
    w.document.close();
}}

function download(content, filename, type) {{
    const blob = new Blob([content], {{ type: type }});
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = filename;
    a.click();
}}

// 页面加载时恢复之前保存的数据
window.onload = function() {{
    const saved = localStorage.getItem('7m_id_map');
    if (saved) {{
        const data = JSON.parse(saved);
        for (const [idx, id] of Object.entries(data)) {{
            const input = document.getElementById('id_' + idx);
            if (input) input.value = id;
        }}
        updateProgress();
    }}
}};

// 自动保存
setInterval(() => {{
    const data = {{}};
    for (let i = 0; i < {len(players)}; i++) {{
        const val = document.getElementById('id_' + i).value.trim();
        if (val) data[i] = val;
    }}
    localStorage.setItem('7m_id_map', JSON.stringify(data));
}}, 2000);
</script>

</body>
</html>'''

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'[完成] HTML 助手已生成: {output_path}')
    print(f'  用浏览器打开: file://{os.path.abspath(output_path)}')
    print(f'  填写完后点击"导出 CSV"，然后运行:')
    print(f'  python fifa_build_id_map.py star_players_7m_id.csv')


if __name__ == '__main__':
    players = read_star_players(EXCEL_PATH)
    print(f'[信息] 共读取 {len(players)} 名焦点球星')
    generate_html(players, OUTPUT_HTML)
