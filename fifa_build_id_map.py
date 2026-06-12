#!/usr/bin/env python3
"""
2026世界杯球员 7m ID 自动匹配 + 批量抓取
用法:
  python fifa_build_id_map.py                # 生成填写模板
  python fifa_build_id_map.py --search       # 自动搜索匹配7m ID（需联网）
  python fifa_build_id_map.py filled.csv    # 根据填好的CSV批量抓取并导出Excel
"""

import json
import re
import sys
import time
import csv
import os
import argparse
import urllib.request
import urllib.parse

TEMPLATE_CSV = 'player_id_template.csv'
OUTPUT_XLSX = '2026世界杯球员数据_7m.xlsx'
EXCEL_PATH = '2026世界杯观赛指南.xlsx'


# ─── 读取Excel球员名单 ───────────────────────────────────────────────────────

def read_player_list(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        print('[错误] 需要 openpyxl，请运行: pip install openpyxl', file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if '球员名单' not in wb.sheetnames:
        print('[错误] Excel中找不到"球员名单"Sheet', file=sys.stderr)
        sys.exit(1)

    ws = wb['球员名单']
    players = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=4).value  # 姓名
        if not name:
            continue
        players.append({
            'group': _cell(ws, row, 1),
            'team': _cell(ws, row, 2),
            'number': _cell(ws, row, 3),
            'name_cn': str(name).strip(),
            'name_en': _cell(ws, row, 5),
            'position': _cell(ws, row, 6),
            'birthday': _cell(ws, row, 8),
        })
    return players


def _cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ''
    return str(v).strip()


# ─── 生成手动填写模板 ─────────────────────────────────────────────────────────

def generate_template(players, csv_path):
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['分组', '队伍', '号码', '姓名', '英文名', '位置', '出生日期', '7m球员ID', '备注'])
        for p in players:
            writer.writerow([
                p['group'], p['team'], p['number'],
                p['name_cn'], p['name_en'], p['position'], p['birthday'],
                '', ''
            ])
    print(f'[完成] 模板已生成: {csv_path}')
    print(f'  请填写"7m球员ID"列，保存后运行:')
    print(f'  python {os.path.basename(__file__)} {csv_path}')
    return csv_path


# ─── 通过 DuckDuckGo 搜索 7m 球员ID ──────────────────────────────────────

def search_7m_id_via_ddg(player_name, lang='zh'):
    """
    通过 DuckDuckGo HTML 搜索 site:player.7m.com.cn <球员名>
    返回 [(id, url)] 列表，最多返回3个结果。
    """
    query = f'site:player.7m.com.cn {player_name}'
    params = urllib.parse.urlencode({'q': query, 'kl': 'zh-cn', 'df': 'd'})
    url = f'https://html.duckduckgo.com/html/?{params}'

    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept-Language': 'zh-CN,zh;q=0.9',
        })
        with urllib.request.urlopen(req, timeout=10) as r:
            html = r.read().decode('utf-8', errors='ignore')

        # 提取 player.7m.com.cn/数字/ 形式的链接
        matches = re.findall(
            r'player\.7m\.com\.cn/(\d+)/',
            html
        )
        # 去重并保持顺序
        seen = set()
        results = []
        for mid in matches:
            if mid not in seen:
                seen.add(mid)
                results.append(mid)
                if len(results) >= 3:
                    break
        return results
    except Exception as e:
        return []


def search_7m_id_direct(player_name):
    """
    直接访问 7m 搜索页，解析结果中的球员链接。
    """
    query = urllib.parse.quote(player_name)
    # 尝试简体中文搜索页
    for search_url in [
        f'https://search.7m.com.cn/search.aspx?q={query}&type=player&lang=gb',
        f'https://www.7m.com.cn/search.aspx?q={query}&type=player',
    ]:
        try:
            req = urllib.request.Request(search_url, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
                'Referer': 'https://www.7m.com.cn/',
            })
            with urllib.request.urlopen(req, timeout=10) as r:
                html = r.read().decode('utf-8', errors='ignore')
            ids = re.findall(r'player\.7m\.com\.cn/(\d+)/', html)
            if ids:
                return list(dict.fromkeys(ids))[:3]  # 去重
        except Exception:
            continue
    return []


def auto_match_ids(players, csv_path, use_ddg=True):
    """
    自动搜索匹配 7m ID，输出到 CSV（含匹配状态）。
    use_ddg: 是否使用 DuckDuckGo 搜索（更准确但较慢）
    """
    print(f'[自动匹配] 共 {len(players)} 名球员，开始搜索 7m ID...')
    print('  （这可能需要几分钟，请耐心等待）')

    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['分组', '队伍', '号码', '姓名', '英文名', '位置', '出生日期',
                         '7m球员ID', '匹配方式', '备注'])

        for i, p in enumerate(players):
            name = p['name_cn']
            print(f'  [{i+1}/{len(players)}] {name} ({p["team"]})',
                  end=' ', file=sys.stderr)

            ids = []
            if use_ddg:
                ids = search_7m_id_via_ddg(name)
                if not ids:
                    ids = search_7m_id_via_ddg(p.get('name_en', ''))
            if not ids:
                ids = search_7m_id_direct(name)
                if not ids and p.get('name_en'):
                    ids = search_7m_id_direct(p['name_en'])

            if len(ids) == 1:
                print(f'-> {ids[0]} (唯一匹配)', file=sys.stderr)
                writer.writerow([
                    p['group'], p['team'], p['number'], name,
                    p['name_en'], p['position'], p['birthday'],
                    ids[0], '自动匹配', ''
                ])
            elif len(ids) > 1:
                print(f'-> {ids[0]} (共{len(ids)}个，需复核)', file=sys.stderr)
                writer.writerow([
                    p['group'], p['team'], p['number'], name,
                    p['name_en'], p['position'], p['birthday'],
                    ids[0], '需复核', f'搜索到{len(ids)}个结果: {",".join(ids)}'
                ])
            else:
                print('-> 未找到', file=sys.stderr)
                writer.writerow([
                    p['group'], p['team'], p['number'], name,
                    p['name_en'], p['position'], p['birthday'],
                    '', '未找到', '请手动查找'
                ])

            # 避免请求过快被封
            if use_ddg:
                time.sleep(1.5)
            else:
                time.sleep(0.8)

    print(f'\n[完成] 搜索结果已保存: {csv_path}')
    print(f'  请打开CSV检查"7m球员ID"列，修正错误匹配后保存。')
    print(f'  确认无误后运行: python {os.path.basename(__file__)} {csv_path}')


# ─── 读取填好的CSV ───────────────────────────────────────────────────────────

def load_filled_csv(csv_path):
    """读取填好7m ID的CSV，返回球员列表"""
    players = []
    with open(csv_path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            pid = row.get('7m球员ID', '').strip()
            if pid and pid.isdigit():
                players.append({
                    'group': row.get('分组', ''),
                    'team': row.get('队伍', ''),
                    'number': row.get('号码', ''),
                    'name_cn': row.get('姓名', ''),
                    'name_en': row.get('英文名', ''),
                    '7m_id': pid,
                })
    return players


# ─── 批量抓取并导出Excel ────────────────────────────────────────────────────

def batch_scrape_and_export(players, output_xlsx):
    """调用 scrape_7m_player.py 批量抓取，然后合并导出Excel"""
    import subprocess

    if not players:
        print('[错误] 没有有效的球员ID', file=sys.stderr)
        sys.exit(1)

    print(f'[抓取] 共 {len(players)} 名球员')
    ids = [p['7m_id'] for p in players]

    # 调用 scrape_7m_player.py
    script_dir = os.path.dirname(os.path.abspath(__file__)) or '.'
    scrape_py = os.path.join(script_dir, 'scrape_7m_player.py')
    tmp_json = os.path.join(script_dir, '_tmp_players.json')

    cmd = [sys.executable, scrape_py] + ids + ['-o', tmp_json]
    print(f'[执行] {" ".join(cmd)}', file=sys.stderr)
    ret = subprocess.run(cmd, cwd=script_dir)

    if ret.returncode != 0 or not os.path.exists(tmp_json):
        print('[错误] 抓取失败', file=sys.stderr)
        sys.exit(1)

    # 读取JSON结果
    with open(tmp_json, 'r', encoding='utf-8') as f:
        data = json.load(f)

    results = data if isinstance(data, list) else [data]

    # 导出Excel（调用 scrape_7m_player 的 export_to_xlsx 函数）
    import importlib.util
    spec = importlib.util.spec_from_file_location('scrape_mod', scrape_py)
    scrape_mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(scrape_mod)

    scrape_mod.export_to_xlsx(results, output_xlsx)
    os.remove(tmp_json)
    print(f'[完成] Excel已保存: {output_xlsx}')


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='2026世界杯球员 7m ID匹配 + 批量抓取')
    parser.add_argument('csv', nargs='?', help='填好7m ID的CSV文件')
    parser.add_argument('--search', action='store_true',
                        help='自动搜索匹配7m ID（生成CSV供复核）')
    parser.add_argument('--no-ddg', action='store_true',
                        help='不使用DuckDuckGo搜索（仅用7m搜索，较快但不稳定）')
    parser.add_argument('--xlsx', default=EXCEL_PATH,
                        help=f'Excel文件路径（默认: {EXCEL_PATH}）')
    parser.add_argument('--output', '-o', default=OUTPUT_XLSX,
                        help=f'输出Excel文件（默认: {OUTPUT_XLSX}）')
    args = parser.parse_args()

    # 没有CSV也没有--search → 生成模板
    if not args.csv and not args.search:
        if not os.path.exists(args.xlsx):
            print(f'[错误] 找不到 {args.xlsx}', file=sys.stderr)
            sys.exit(1)
        players = read_player_list(args.xlsx)
        if not players:
            print('[错误] 无法读取球员名单', file=sys.stderr)
            sys.exit(1)
        print(f'[信息] 共读取 {len(players)} 名球员')
        generate_template(players, TEMPLATE_CSV)
        print()
        print('[提示] 如何查找 7m 球员ID：')
        print('  方法1（推荐）: 运行自动搜索:')
        print(f'    python {os.path.basename(__file__)} --search')
        print('  方法2: 手动查找:')
        print('    1. 浏览器打开 https://player.7m.com.cn/')
        print('    2. 搜索球员中文名或英文名')
        print('    3. 打开球员页面，URL中 player.7m.com.cn/XXXXXX/ 的 XXXXXX 即球员ID')
        return

    # --search 模式 → 自动搜索
    if args.search:
        if not os.path.exists(args.xlsx):
            print(f'[错误] 找不到 {args.xlsx}', file=sys.stderr)
            sys.exit(1)
        players = read_player_list(args.xlsx)
        auto_match_ids(players, TEMPLATE_CSV, use_ddg=not args.no_ddg)
        return

    # 有CSV → 批量抓取
    if not os.path.exists(args.csv):
        print(f'[错误] 找不到 {args.csv}', file=sys.stderr)
        sys.exit(1)

    players = load_filled_csv(args.csv)
    if not players:
        print('[错误] CSV中没有有效的7m球员ID', file=sys.stderr)
        sys.exit(1)

    batch_scrape_and_export(players, args.output)


if __name__ == '__main__':
    main()
